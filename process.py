from openpyxl import Workbook
from openpyxl import load_workbook
from scipy.stats import iqr
from datetime import date
import numpy
import csv
import re

integer_prog = re.compile('([\d]+)');
float_prog = re.compile('([0-9]*\.[0-9]+|[0-9]+)');
off_prog = re.compile('(Off)|(off)');
on_prog = re.compile('(On)|(on)');
workbook_in = load_workbook(filename="osu! Top Mouse Player List.xlsx")

num_values = 0
aggregate = 0
values = []

OS_MULTIPLIER = {"1":1/32.0, "2":1/16.0, "3":1/4.0, "4":1/2.0, "5":3/4.0, "6":1.0, "7":1.5, "8":2.0, "9":2.5, "10":3.0, "11":3.5};
DPI = 6
OS = 7
MULTIPLIER = 8
RESOLUTION = 9
RAW = 10
FILE_NAME = "osu_sens_{:0>2d}-{:0>2d}-{:0>2d}.txt".format(date.today().month, date.today().day, date.today().year)

for i in range (4, 640):

    sheet = workbook_in.active
    dpi = str(sheet.cell(row = i, column = DPI).value).replace("dpi", "")
    
    if (dpi == "None"):
        continue

    dpi_result = integer_prog.match(dpi)
    if (dpi_result):
        dpi = dpi_result.group(0)

    try:
        dpi_val = float(dpi)
    except ValueError:
        print("dpi error at: " + str(i))
        print("value of: " + dpi)
        continue

    multiplier = str(sheet.cell(row = i, column = MULTIPLIER).value).replace("x", "")

    if (multiplier == "None"):
            continue

    multiplier_result = float_prog.match(multiplier)
    if (multiplier_result):
        multiplier = multiplier_result.group(0)

    try:
        multiplier_val = float(multiplier)
    except ValueError:
        print("multiplier error at: " + str(i))
        print("value of: " + multiplier)
        continue

    os = str(sheet.cell(row = i, column = OS).value)

    if (os == "None"):
        continue
    if (re.search("(On)|(on)", os)):
        continue
    os = re.search("(\d+)(?=/11)", os)
    os = os.group(0)

    os_multiplier = OS_MULTIPLIER[os]
    

    raw = str(sheet.cell(row = i, column = RAW).value)
    if (re.search("(On)|(on)", raw)) or not (re.search("(Off)|(off)", raw)) :
        os_multiplier = 1.0

    resolution = str(sheet.cell(row = i, column = RESOLUTION).value)
    resolution_width_search = re.search("(?P<value>\d+)(?=x\d+)", resolution)
    resolution_height_search = re.search("(\d+x)(?P<value>\d+)", resolution)
    if (resolution_width_search == None or resolution_height_search == None):
        continue

    try:
        text = resolution_width_search.group('value')
        resolution_width = float(text)
    except ValueError:
        print("resolution_width error at: " + str(i))
        print("value of: " + text)
        continue
    try:
        text = resolution_height_search.group('value')
        resolution_height = float(resolution_height_search.group('value'))
    except ValueError:
        print("resolution_height error at: " + str(i))
        print("value of: " + text)
        continue

    # pixel height is assumed to control the Osu play area
    # pixel height is assumed to control the Osu play area linearly
    # default pixel height is assumed to be 1080. If the user uses a different resolution, then this needs to be changed for equivalent results
    #   i.e. if someone used 400 dpi at a resolution height of 1000, this is assumed to be the equivalent to using 800dpi at a height of 2000
    USER_RESOLUTION_HEIGHT = 1080

    pixel_height_multiplier = USER_RESOLUTION_HEIGHT/resolution_height
    this_value = dpi_val*multiplier_val*os_multiplier*pixel_height_multiplier
    values.append(this_value)
    aggregate = aggregate + this_value
    num_values = num_values + 1

np_values = numpy.array(values)
q1 = numpy.quantile(np_values, .25)
q2 = numpy.quantile(np_values, .5)
q3 = numpy.quantile(np_values, .75)
av = aggregate/num_values

file_content = ""
file_content += "Q1 " + str(q1) + "\r\n"
file_content += "Q2 " + str(q2) + "\r\n"
file_content += "Q3 " + str(q3) + "\r\n"
file_content += "AV " + str(av)

f = open(FILE_NAME, "w")
f.write(file_content)
f.close()

print("finished processing")
